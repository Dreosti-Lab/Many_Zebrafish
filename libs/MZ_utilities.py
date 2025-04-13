# -*- coding: utf-8 -*-
"""
Many_Zebrafish: Utility Library

@author: kampff
"""
# Import libraries
import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import glob
import shutil
import cv2
import re
from datetime import datetime, timedelta
import zipfile
from io import StringIO
import pandas as pd

# Utilities for analysing 96-well plate experiments

# Extract PPI stimulus from LED intensity time series
def extract_ppi_stimuli(led_intensity):
    baseline = np.median(led_intensity)
    signal = led_intensity - baseline
    threshold = np.max(signal) / 4
    pulses = []
    last_peak = 0
    for i, s in enumerate(signal):
        if s > threshold:
            if (i - last_peak) > 3:
                pulses.append(i)
                last_peak = i
    single_pulses = []
    paired_pulses = []
    p = 0
    while p < len(pulses):
        if (p+1) == len(pulses):
            single_pulses.append(pulses[p])
            break
        if (pulses[p+1] - pulses[p]) > 1000:
            single_pulses.append(pulses[p])
            p = p + 1
        else:
            paired_pulses.append((pulses[p], pulses[p+1]))
            p = p + 2
            
    return single_pulses, paired_pulses

# Extract behaviour for a well (fish) from a plate
def extract_behaviour(plate, well):
    num_frames = len(plate.intensity)
    behaviour = np.zeros((num_frames, 5), dtype=np.float32)
    behaviour[:,0] = plate.wells[well].x
    behaviour[:,1] = plate.wells[well].y
    behaviour[:,2] = plate.wells[well].heading
    behaviour[:,3] = plate.wells[well].area
    behaviour[:,4] = plate.wells[well].motion
    return behaviour

# Extract response around a frame index (e. g. stimulus or PPI pulse)
def extract_response(behaviour, index_frame, pre_frames, post_frames):
        response = np.zeros((pre_frames+post_frames+1, 5))
        response = behaviour[(index_frame-pre_frames):(index_frame+post_frames+1),:]
        return response

# Extract responses around a list of frame indices
def extract_responses(behaviour, index_frames, pre_frames, post_frames):
    responses = np.zeros((pre_frames+post_frames+1, 5, len(index_frames)))
    for i, index_frame in enumerate(index_frames):
        responses[:, :, i] = extract_response(behaviour, index_frame, pre_frames, post_frames)
    return responses

# Load path list
def load_path_list(path_list_path):
    tmp_path_list = open(path_list_path,'r').read().split('\n')
    path_list = []
    for path in tmp_path_list:
        if path != '':
            path_list.append(path)
    return path_list

# Parse summary (PPI)
def parse_summary_PPI(summary_path, gene_name):
    first_row = 2
    last_row = 3841

    # Load summary workbook (PPI)
    wb = load_workbook(summary_path, read_only=True)
    ws = wb.get_sheet_by_name('Ppi')

    # Extract cells
    plate_cells = ws[f'A{first_row}:A{last_row}']
    path_cells = ws[f'F{first_row}:F{last_row}']
    well_cells = ws[f'J{first_row}:J{last_row}']
    include_cells = ws[f'K{first_row}:K{last_row}']
    control_cells = ws[f'M{first_row}:M{last_row}']
    gene_cells = ws[f'N{first_row}:N{last_row}']

    # Find plates
    plates = []
    paths = []
    last_plate = -1
    for i, cell in enumerate(gene_cells):
        gene = cell[0].value
        if gene == gene_name:
            plate = plate_cells[i][0].value
            path = '/' + path_cells[i][0].value
            if plate != last_plate:
                plates.append(plate)
                paths.append(path)
                last_plate = plate

    # Extract valid controls and test fish
    all_controls = []
    all_tests = []
    for plate in plates:
        controls = []
        tests = []
        for i, cell in enumerate(plate_cells):
            if cell[0].value != plate:                  # Correct plate
                continue
            if include_cells[i][0].value == 0:          # Include?
                continue
            if gene_cells[i][0].value == gene_name:     # Correct gene?
                tests.append(well_cells[i][0].value)
            elif control_cells[i][0].value == 1:        # Is control?
                controls.append(well_cells[i][0].value)
            else:
                continue
        all_controls.append(controls)
        all_tests.append(tests)
    return plates, paths, all_controls, all_tests

# Parse summary (Sleep)
def parse_summary_Sleep(summary_path, gene_name):
    first_row = 2
    last_row = 3361

    # Load summary workbook (Sleep)
    wb = load_workbook(summary_path, read_only=True)
    ws = wb.get_sheet_by_name('Sleep')

    # Extract cells
    plate_cells = ws[f'A{first_row}:A{last_row}']
    path_cells = ws[f'F{first_row}:F{last_row}']
    well_cells = ws[f'I{first_row}:I{last_row}']
    include_cells = ws[f'J{first_row}:J{last_row}']
    control_cells = ws[f'L{first_row}:L{last_row}']
    gene_cells = ws[f'M{first_row}:M{last_row}']

    # Find plates
    plates = []
    paths = []
    last_plate = -1
    for i, cell in enumerate(gene_cells):
        gene = cell[0].value
        if gene == gene_name:
            plate = plate_cells[i][0].value
            path = '/' + path_cells[i][0].value
            if plate != last_plate:
                plates.append(plate)
                paths.append(path)
                last_plate = plate

    # Extract valid controls and test fish
    all_controls = []
    all_tests = []
    for plate in plates:
        controls = []
        tests = []
        for i, cell in enumerate(plate_cells):
            if cell[0].value != plate:                  # Correct plate
                continue
            if include_cells[i][0].value == 0:          # Include?
                continue
            if gene_cells[i][0].value == gene_name:     # Correct gene?
                tests.append(well_cells[i][0].value)
            elif control_cells[i][0].value == 1:        # Is control?
                controls.append(well_cells[i][0].value)
            else:
                continue
        all_controls.append(controls)
        all_tests.append(tests)
    return plates, paths, all_controls, all_tests

# Create a folder (if it does not exist)
def create_folder(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return

# Clear a folder (or create)
def clear_folder(folder_path):
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)
    return

# Extract experiment start time
def get_start_date_time(phc_path):
    # Read first line of PHC file
    with open(phc_path, 'r', encoding='utf-8') as f:
        line = f.readline()

    # Extract RunDate and RunTime using regex
    match = re.search(r'RunDate="([\d/]+)".*?RunTime="([\d:]+ [APM]+)"', line)
    if match:
        date_str = match.group(1)  # e.g., '2022/09/19'
        time_str = match.group(2)  # e.g., '7:09:55 PM'
        
        # Combine and parse into datetime
        combined_str = f"{date_str} {time_str}"
        dt = datetime.strptime(combined_str, "%Y/%m/%d %I:%M:%S %p")
        return dt
    else:
        raise ValueError("RunDate and RunTime not found in the first line.")

# Extract ordered frame times
def get_ordered_frame_times(zip_path):
    # Initialize sets to store unique timestamps per plate
    timestamps_A = set()
    timestamps_B = set()

    # Open the archive
    archive = zipfile.ZipFile(zip_path, 'r')
    xls_files = [zf for zf in archive.filelist if zf.filename.lower().endswith(('.xls', '.xlsx'))]

    # Loop through all XLS files
    for i, zf in enumerate(xls_files):
        try:
            raw = archive.read(zf.filename).decode('ascii')
            df = pd.read_csv(StringIO(raw), sep='\t')
        except Exception as e:
            print(f"Error reading {zf.filename}: {e}")
            continue

        # Filter rows with numeric locations and timestamps
        df = df[df['location'].str.startswith('C') & df['abstime'].apply(lambda x: str(x).isdigit())]

        # Extract location numbers
        df['loc_num'] = df['location'].str[1:].astype(int)
        df['ts'] = df['abstime'].astype(int)

        # Apply masks for location ranges
        mask_A = df['loc_num'].between(1, 96)
        mask_B = df['loc_num'].between(97, 192)

        # Add timestamps to corresponding sets
        timestamps_A.update(df.loc[mask_A, 'ts'])
        timestamps_B.update(df.loc[mask_B, 'ts'])

        # Report
        print(f"Processed {i}/{len(xls_files)}: {zf.filename}")

    # Convert to sorted lists
    sorted_A = sorted(timestamps_A)
    sorted_B = sorted(timestamps_B)

    return sorted_A, sorted_B

# Extract sunset and sunrise frame indices
def get_sunset_sunrise_frames(start_datetime, frametimes):
    MICROSECONDS_IN_DAY = 24 * 60 * 60 * 1_000_000
    MICROSECONDS_IN_10_HOURS = 10 * 60 * 60 * 1_000_000

    # Compute frame time offsets
    frame_offsets = np.array(frametimes) - frametimes[0]

    # Compute first sunset (11 PM on the same day, or next day if already past)
    first_sunset_dt = start_datetime.replace(hour=23, minute=0, second=0, microsecond=0)
    first_sunset_us = int((first_sunset_dt - start_datetime).total_seconds() * 1e6)
    if start_datetime >= first_sunset_dt:
        print("Experiment Flaw: started after 11 pm")
        first_sunset_dt += timedelta(days=1)

    # List sunset and sunrise offset times
    sunset_offsets = []
    sunrise_offsets = []
    max_us = frametimes[-1]
    current_sunset_us = first_sunset_us
    while current_sunset_us <= max_us:
        sunset_offsets.append(current_sunset_us)
        sunrise_offsets.append(current_sunset_us + MICROSECONDS_IN_10_HOURS)  # sunrise = 10 hrs after sunset
        current_sunset_us += MICROSECONDS_IN_DAY
    print(sunset_offsets)
    print(sunrise_offsets)

    # Use searchsorted to get frame indices
    sunset_indices = np.searchsorted(frame_offsets, sunset_offsets, side='right')
    sunrise_indices = np.searchsorted(frame_offsets, sunrise_offsets, side='right')

    return sunset_indices, sunrise_indices

#FIN